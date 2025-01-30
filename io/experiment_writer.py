import datetime
import os
from typing import Any

import openpyxl
import pandas as pd

from rent_buy_invest.utils import io_utils


class ExperimentWriter:
    """Handles writing outputs to an output directory for a single experiment.

    Attributes:
        _output_dir (str): The output directory for the experiment"""

    DEFAULT_OUTPUT_DIR_PROJECT_PATH: str = "rent_buy_invest/out/"

    def __init__(
        self, experiment_name: str, output_dir_project_path: str | None = None
    ) -> None:
        """Initialize ExperimentWriter

        Args:
            experiment_name: Name of the experiment; it must contain only alphanumeric characters, "_", and "-"
            output_dir_project_path: Optional output dir path relative to folder containing rent_buy_invest.
                If not provided, ExperimentWriter.DEFAULT_OUTPUT_DIR_PROJECT_PATH is used
        """
        assert all(
            c.isalpha() or c.isdigit() or c in "_-" for c in experiment_name
        ), f"Provide an experiment name which only contains characters, digits, underscores, and/or dashes; received '{experiment_name}'"
        timestamp_str = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        output_dir_project_path = (
            output_dir_project_path
            if output_dir_project_path
            else ExperimentWriter.DEFAULT_OUTPUT_DIR_PROJECT_PATH
        )
        self._output_dir = os.path.join(
            ExperimentWriter.DEFAULT_OUTPUT_DIR_PROJECT_PATH,
            experiment_name,
            timestamp_str,
        )
        io_utils.make_dirs(self._output_dir)

    def write_output_yaml(self, filename: str, obj: Any) -> None:
        path = os.path.join(self._output_dir, filename)
        io_utils.write_yaml(path, obj)

    def write_output_xlsx_df(
        self, filename: str, df: pd.DataFrame, num_header_rows=0
    ) -> None:
        # TODO move some of this to io_utils xlsx writing method... and make some of this parametrizable or auto-set
        path = os.path.join(self._output_dir, filename)
        io_utils.write_xlsx_df(path, df)
        wb = openpyxl.load_workbook(path)
        ws = wb["Sheet1"]
        ws.column_dimensions["A"].width = 15
        ws.freeze_panes = f"B{num_header_rows+1}"
        max_num_cols_with_data = 14
        for i in range(max_num_cols_with_data):
            col_name = chr(ord("B") + i)
            # ws.column_dimensions[col_name].number_format = "$#,##0.00"
            ws.column_dimensions[col_name].width = 18
            for header_row in range(1, num_header_rows + 1):
                ws[f"{col_name}{header_row}"].alignment = openpyxl.styles.Alignment(
                    horizontal="left"
                )
            for cell in ws[col_name]:
                cell.number_format = "$#,##0.00"
        wb.save(path)
